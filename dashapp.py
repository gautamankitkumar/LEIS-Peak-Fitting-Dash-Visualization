# -*- coding: utf-8 -*-
import dash
import dash_core_components as dcc
import dash_html_components as html
from dash.dependencies import Input, Output
import numpy as np
import pandas as pd
import ujson
# Used to read excel file
from openpyxl import load_workbook
import plotly.graph_objects as go


excel_file = '500 K LEIS 1st round 27.5.xlsx'
def read_data(filename):    
    """ 
    This function reads the above specified file and returns two numpy array
    which contain the x-data and y-data 

    filename: string, Excel file to be read

    """
    
    # Iterate over all X-values. Y-values are stored in colummns of particular worksheet
    for x in range(0,13):

        wb = load_workbook(filename)
        ws = wb[str(x)]

        # This position of metadata doesn't change its relative position from sheet-to-sheet
        n_energy = ws['D2'].value
        n_iter = ws['D5'].value
        Rows_to_Skip = 15

        # Rename columns
        column_names = [str(x) for x in range(0,n_iter)]
        column_names.insert(0,'nan')
        column_names.insert(0,'KE')

        # Read data using pandas
        df_data = pd.read_excel(io = filename,
                           sheet_name=x,
                           skiprows = Rows_to_Skip,
                           names = column_names,
                           index_col='KE'
                          )
        # Drop the second column as it is always supposed to be false
        df_data.drop(columns=df_data.columns[0],inplace=True)
        
        # Get x_data as the index 
        x_array = np.array(df_data.index).reshape(len(df_data.index),1)
        
        # If we encounter first sheet
        if x==0:
            y = df_data.to_numpy()
            
        # Stack with the cummulative y built till now
        else:
            y = np.hstack((y, df_data.to_numpy()))
            
    # Ideally x_array should be (481, 1), and y should be (481, 169)
    return x_array, y


import pickle
# These variables are used subsequently in the code
# x_data,y_data = read_data(excel_file)
# print(x_data.shape,y_data.shape)

# with open('x_data', 'wb') as f:
#     # Pickle the 'data' dictionary using the highest protocol available.
#     pickle.dump(x_data, f, pickle.HIGHEST_PROTOCOL)

# with open('y_data', 'wb') as f:
#     # Pickle the 'data' dictionary using the highest protocol available.
#     pickle.dump(y_data, f, pickle.HIGHEST_PROTOCOL)

with open('x_data', 'rb') as f:
    # The protocol version used is detected automatically, so we do not
    # have to specify it.
    x_data = pickle.load(f)

with open('y_data', 'rb') as f:
    # The protocol version used is detected automatically, so we do not
    # have to specify it.
    y_data = pickle.load(f)

with open('data.json') as json_file:
    peak_data = ujson.load(json_file)

app = dash.Dash(__name__)

grid_points = []
for i in range(169):
    grid_points.append({'label':str(i//13)+', '+str(i%13),'value':i})


app.layout = html.Div(children=[
    html.H2(children="Pick an index from the left graph \n to show the data on the right graph"),
    dcc.Checklist(id='show-fit',options=[{'label':'Show Fitted lines','value':0}]),
dcc.Checklist(id='index-grid',
    options=grid_points,
    value=[1]
),
dcc.Graph(id='ternary-plot'),
html.Button('Clear All',id='clear-button',title='Clears all the data graphs from the window', n_clicks=1),
dcc.Graph(id='data-graph')
])


@app.callback(
    Output('data-graph','figure'),
    [Input('index-grid','value'),
    Input('show-fit','value')]
)
def update_graph(in1,show_fit):

    if not in1:
        return {'data':[]}
    traces = []
    if show_fit:        
        for val in in1:
            yfit = peak_data[str(val)]['fit']
            legendgroup_name = 'group'+str(val)

            traces.append(dict(
            x=np.array(x_data)[:,0].tolist(),
            y=np.array(y_data)[:,val].tolist(),
            text=str(val),
            mode='markers',
            opacity=0.5,
            legendgroup=legendgroup_name,
            name='Data PointsX:'+str(val//13) + ', Y:' + str(val%13),
            marker = dict(color=str(val)),
                ))
            traces.append(dict(
            x=np.array(x_data)[:,0].tolist(),
            y=yfit,
            text=str(val),
            mode='line',
            opacity=0.7,
            legendgroup=legendgroup_name,
            name='Fitted, X:'+str(val//13) + ', Y:' + str(val%13),
            line = dict(color=str(val)),
                ))
    else:        
        for val in in1:
            traces.append(dict(
            x=np.array(x_data)[:,0].tolist(),
            y=np.array(y_data)[:,val].tolist(),
            text=str(val),
            mode='markers',
            opacity=0.7,
            name='Data Points X:'+str(val//13) + ', Y:' + str(val%13),
                ))        

    return {
        'data': traces,
        'layout': dict(
            xaxis={ 'title': 'KE'},
            yaxis={'title': 'Data points'},
            margin={'l': 40, 'b': 40, 't': 10, 'r': 10},
            legend={'x': 0, 'y': 1},
            hovermode='closest',
        )
    }

current_n = 0
@app.callback(
Output('index-grid','value'),
[Input('clear-button','n_clicks')]
)
def clear_graph(n_click):
    global current_n
    if n_click>current_n:
        current_n = n_click
        return []

@app.callback(
    Output('ternary-plot','figure'),
    [Input('index-grid','value')]
)


def get_ternary_plot(input_values):
    if not input_values:
        return {'data':[]}
    traces = []
    metal_1 = []
    metal_2 = []
    metal_3 = []
    for val in input_values:
        metal_1.append(val//13)
        metal_2.append(val%13)
    fig =  go.Figure(go.Scatterternary(text='Metal Composition',a=metal_1,b=metal_2,
    mode='markers',marker={'symbol': 100,'size': 10},))
    fig.update_layout({
    'title': 'Index Grid Points on CSAF',
    'ternary': {
        'sum': 39,
    'aaxis':{'title':'Component - A'},
    'baxis':{'title':'Component - B'},
    'caxis':{'title':'Component - C'},
    }})

    return fig

if __name__ == '__main__':
    app.run_server(debug=True)